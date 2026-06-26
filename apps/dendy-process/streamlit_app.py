from io import BytesIO
import re

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import streamlit as st


st.set_page_config(page_title="Dendy Process", layout="wide")
st.title("Dendy Process")


def load_workbook_bundle(file_bytes: bytes):
	workbook = load_workbook(BytesIO(file_bytes))
	sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine="openpyxl")
	return workbook, sheets


def load_first_sheet(uploaded_file) -> pd.DataFrame:
	uploaded_file.seek(0)
	return pd.read_excel(uploaded_file, engine="openpyxl")


def load_sheet_matrix(uploaded_file) -> pd.DataFrame:
	uploaded_file.seek(0)
	return pd.read_excel(uploaded_file, header=None, engine="openpyxl")


def normalize_site_value(value):
	if value is None:
		return None
	text = str(value).strip()
	if text == "" or text.lower() == "nan":
		return None
	try:
		f = float(text)
		if f.is_integer():
			return str(int(f))
	except Exception:
		pass
	return text.upper()



def parse_site_rep(site_rep_value):
	text = str(site_rep_value).strip().upper()
	if text == "" or text == "NAN":
		return None, None

	# General pattern: trailing letter is always replicate.
	# Examples: 6A -> site 6 rep 1, GILLA -> site GILL rep 1.
	m = re.match(r"^(.+?)([A-Z])$", text)
	if m:
		site_text = m.group(1).strip()
		site = normalize_site_value(site_text)
		rep = ord(m.group(2)) - ord("A") + 1
		return site, rep

	# Fallback pattern: 6,1 or 6-1
	m = re.match(r"^(\d+)\s*[,\-/]\s*(\d+)$", text)
	if m:
		return normalize_site_value(m.group(1)), int(m.group(2))

	# Fallback: plain site number with unknown rep
	m = re.match(r"^(\d+)$", text)
	if m:
		return normalize_site_value(m.group(1)), None

	return normalize_site_value(text), None


def normalize_header_token(value) -> str:
	text = str(value).strip().lower()
	text = re.sub(r"[^a-z0-9]+", "_", text)
	return text.strip("_")


def find_sheet_case_insensitive(workbook, sheet_name: str):
	target = sheet_name.strip().lower()
	for ws in workbook.worksheets:
		if ws.title.strip().lower() == target:
			return ws
	return None


def find_column_by_aliases(raw_matrix: pd.DataFrame, row_candidates: list[int], aliases: set[str]):
	for row_idx in row_candidates:
		if row_idx < 0 or row_idx >= len(raw_matrix):
			continue
		for col_idx, value in enumerate(raw_matrix.iloc[row_idx].tolist()):
			if normalize_header_token(value) in aliases:
				return col_idx
	raise ValueError(f"Could not find required column with aliases: {sorted(aliases)}")


def build_data_manager_export(raw_matrix: pd.DataFrame, export_year: int, data_start_col_1_based: int = 12):
	data_start_col = data_start_col_1_based - 1

	# Fixed format coordinates from the raw file layout:
	# lake ids start at [1,12], site/rep start at [2,12],
	# taxon data starts at [3:nrow,2], and count data starts at [3,12].
	lake_row = 0
	site_rep_row = 1
	body_start_row = 2
	taxon_col = 1

	if raw_matrix.shape[0] <= body_start_row or raw_matrix.shape[1] <= data_start_col:
		raise ValueError("Raw file is smaller than expected fixed Dendy layout.")

	max_col = raw_matrix.shape[1]
	active_cols = []
	for col_idx in range(data_start_col, max_col):
		lake_val = raw_matrix.iat[lake_row, col_idx]
		site_rep_val = raw_matrix.iat[site_rep_row, col_idx]
		if any(pd.notna(v) and str(v).strip() != "" for v in [lake_val, site_rep_val]):
			active_cols.append(col_idx)

	if not active_cols:
		raise ValueError("No data columns detected at or after the configured data start column.")

	body = raw_matrix.iloc[body_start_row:].copy()
	body = body[body.iloc[:, taxon_col].notna()].copy()
	body = body[body.iloc[:, taxon_col].astype(str).str.strip() != ""].copy()

	counts = body.iloc[:, active_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
	keep_mask = counts.sum(axis=1) > 0
	body_keep = body.loc[keep_mask].copy()
	counts_keep = counts.loc[keep_mask].copy()

	wb = Workbook()
	ws = wb.active
	ws.title = "Sheet1"

	ws.cell(row=1, column=2, value="lakeid")
	ws.cell(row=2, column=2, value="site")
	ws.cell(row=3, column=2, value="rep")
	ws.cell(row=4, column=1, value="taxon_code")
	ws.cell(row=4, column=2, value="year")

	for out_col, src_col in enumerate(active_cols, start=3):
		ws.cell(row=1, column=out_col, value=raw_matrix.iat[lake_row, src_col])
		site_val, rep_val = parse_site_rep(raw_matrix.iat[site_rep_row, src_col])
		ws.cell(row=2, column=out_col, value=site_val)
		ws.cell(row=3, column=out_col, value=rep_val)

	for out_row, (idx, row) in enumerate(body_keep.iterrows(), start=5):
		ws.cell(row=out_row, column=1, value=row.iat[taxon_col])
		ws.cell(row=out_row, column=2, value=export_year)
		for out_col, src_col in enumerate(active_cols, start=3):
			value = counts_keep.at[idx, src_col]
			if pd.notna(value) and float(value) != 0:
				ws.cell(row=out_row, column=out_col, value=int(value) if float(value).is_integer() else float(value))

	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)

	preview_cols = [taxon_col] + active_cols
	preview_df = body_keep.iloc[:, preview_cols].copy()
	preview_df.insert(1, "year", export_year)
	preview_df.columns = ["taxon_code", "year"] + [f"col_{i}" for i in range(1, len(active_cols) + 1)]

	preview_matrix_rows = [
		[None, "lakeid", *[raw_matrix.iat[lake_row, c] for c in active_cols]],
		[None, "site", *[parse_site_rep(raw_matrix.iat[site_rep_row, c])[0] for c in active_cols]],
		[None, "rep", *[parse_site_rep(raw_matrix.iat[site_rep_row, c])[1] for c in active_cols]],
		["taxon_code", "year", *([None] * len(active_cols))],
	]
	for _, row in body_keep.iterrows():
		preview_row = [row.iat[taxon_col], export_year]
		for src_col in active_cols:
			value = counts_keep.at[row.name, src_col]
			if pd.notna(value) and float(value) != 0:
				preview_row.append(int(value) if float(value).is_integer() else float(value))
			else:
				preview_row.append(None)
		preview_matrix_rows.append(preview_row)

	preview_matrix_df = pd.DataFrame(preview_matrix_rows)

	stats = {
		"input_taxa": int(len(body)),
		"kept_taxa": int(len(body_keep)),
		"dropped_taxa": int(len(body) - len(body_keep)),
		"location_columns": int(len(active_cols)),
	}
	return buffer.getvalue(), preview_df, preview_matrix_df, stats


def build_long_table_export(raw_matrix: pd.DataFrame, export_year: int, data_start_col_1_based: int = 12):
	data_start_col = data_start_col_1_based - 1
	lake_row = 0
	site_rep_row = 1
	body_start_row = 2

	if raw_matrix.shape[0] <= body_start_row or raw_matrix.shape[1] <= data_start_col:
		raise ValueError("Raw file is smaller than expected fixed Dendy layout.")

	taxon_col = find_column_by_aliases(raw_matrix, [site_rep_row, lake_row], {"taxon_code", "taxon"})
	family_col = find_column_by_aliases(raw_matrix, [site_rep_row, lake_row], {"family"})
	description_col = find_column_by_aliases(raw_matrix, [site_rep_row, lake_row], {"description", "desc"})
	order_col = find_column_by_aliases(raw_matrix, [site_rep_row, lake_row], {"order_t", "order", "group"})
	class_col = find_column_by_aliases(raw_matrix, [site_rep_row, lake_row], {"class", "class_t"})

	max_col = raw_matrix.shape[1]
	active_cols = []
	for col_idx in range(data_start_col, max_col):
		lake_val = raw_matrix.iat[lake_row, col_idx]
		site_rep_val = raw_matrix.iat[site_rep_row, col_idx]
		if any(pd.notna(v) and str(v).strip() != "" for v in [lake_val, site_rep_val]):
			active_cols.append(col_idx)

	if not active_cols:
		raise ValueError("No data columns detected at or after the configured data start column.")

	body = raw_matrix.iloc[body_start_row:].copy()
	body = body[body.iloc[:, taxon_col].notna()].copy()
	body = body[body.iloc[:, taxon_col].astype(str).str.strip() != ""].copy()

	counts = body.iloc[:, active_cols].apply(pd.to_numeric, errors="coerce").fillna(0)

	records = []
	for idx, row in body.iterrows():
		taxon_code = row.iat[taxon_col]
		family = row.iat[family_col]
		description = row.iat[description_col]
		group_value = row.iat[order_col]
		if pd.isna(group_value) or str(group_value).strip() == "":
			group_value = row.iat[class_col]

		for src_col in active_cols:
			count_value = counts.at[idx, src_col]
			if pd.isna(count_value) or float(count_value) == 0:
				continue

			lakeid = raw_matrix.iat[lake_row, src_col]
			site, rep = parse_site_rep(raw_matrix.iat[site_rep_row, src_col])
			number_indiv = int(count_value) if float(count_value).is_integer() else float(count_value)

			records.append(
				{
					"lakeid": lakeid,
					"year": export_year,
					"site": site,
					"rep": rep,
					"group": group_value,
					"family": family,
					"taxon_code": taxon_code,
					"description": description,
					"number_indiv": number_indiv,
				}
			)

	long_df = pd.DataFrame(
		records,
		columns=[
			"lakeid",
			"year",
			"site",
			"rep",
			"group",
			"family",
			"taxon_code",
			"description",
			"number_indiv",
		],
	)

	buffer = BytesIO()
	long_df.to_excel(buffer, index=False, engine="openpyxl")
	buffer.seek(0)

	stats = {
		"rows": int(len(long_df)),
		"taxa": int(long_df["taxon_code"].nunique()) if not long_df.empty else 0,
		"lakes": int(long_df["lakeid"].nunique()) if not long_df.empty else 0,
	}
	return long_df, buffer.getvalue(), stats


def append_long_table_to_historical(hist_file_bytes: bytes, long_df: pd.DataFrame, sheet_name: str = "data"):
	workbook = load_workbook(BytesIO(hist_file_bytes))
	ws = find_sheet_case_insensitive(workbook, sheet_name)
	if ws is None:
		raise ValueError(f"Sheet '{sheet_name}' was not found in the historical workbook.")

	headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
	header_lookup = {}
	for idx, header in enumerate(headers, start=1):
		token = normalize_header_token(header)
		if token:
			header_lookup[token] = idx

	required_aliases = {
		"lakeid": {"lakeid", "lake_id"},
		"year": {"year", "year4"},
		"site": {"site"},
		"rep": {"rep"},
		"group": {"group"},
		"family": {"family"},
		"taxon_code": {"taxon_code", "taxon"},
		"description": {"description", "desc"},
		"number_indiv": {"number_indiv", "number_individuals", "count", "n"},
	}

	column_map = {}
	for target_col, aliases in required_aliases.items():
		match_col = None
		for alias in aliases:
			if alias in header_lookup:
				match_col = header_lookup[alias]
				break
		if match_col is None:
			raise ValueError(f"Historical data sheet is missing required header for '{target_col}'.")
		column_map[target_col] = match_col

	# Find the true last populated data row (ignoring trailing formatted-but-empty rows).
	data_cols = list(column_map.values())
	last_data_row = 1
	for row_idx in range(ws.max_row, 1, -1):
		if any(ws.cell(row=row_idx, column=col_idx).value not in (None, "") for col_idx in data_cols):
			last_data_row = row_idx
			break

	start_row = last_data_row + 1
	for _, row in long_df.iterrows():
		new_row = start_row
		for col_name, excel_col in column_map.items():
			value = row[col_name]
			if pd.isna(value):
				value = None
			ws.cell(row=new_row, column=excel_col, value=value)
		start_row += 1

	buffer = BytesIO()
	workbook.save(buffer)
	buffer.seek(0)

	stats = {
		"rows_appended": int(len(long_df)),
		"start_row": int(last_data_row + 1),
		"end_row": int(last_data_row + len(long_df)) if len(long_df) > 0 else int(last_data_row + 1),
	}
	return buffer.getvalue(), stats


TARGET_GROUPS = ["Amphipoda", "Diptera", "Ephemeroptera", "Gastropoda", "Odonata", "Oligochaeta"]


def sort_key_part(value):
	if value is None:
		return (2, "")
	text = str(value).strip()
	if text == "":
		return (2, "")
	try:
		return (0, float(text))
	except Exception:
		return (1, text.upper())


def regenerate_site_charts(ws, target_groups: list[str] = TARGET_GROUPS):
	headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
	header_lookup = {}
	for idx, header in enumerate(headers, start=1):
		token = normalize_header_token(header)
		if token:
			header_lookup[token] = idx

	site_col = header_lookup.get("site")
	year_col = header_lookup.get("year4") or header_lookup.get("year")
	if site_col is None or year_col is None:
		return 0

	group_cols = []
	for grp in target_groups:
		grp_col = header_lookup.get(normalize_header_token(grp))
		if grp_col is not None:
			group_cols.append((grp, grp_col))
	if not group_cols:
		return 0

	site_rows = {}
	for row_idx in range(2, ws.max_row + 1):
		site_value = ws.cell(row=row_idx, column=site_col).value
		year_value = ws.cell(row=row_idx, column=year_col).value
		norm_site = normalize_site_value(site_value)
		if norm_site is None or year_value in (None, ""):
			continue

		has_any_value = any(ws.cell(row=row_idx, column=col_idx).value not in (None, "") for _, col_idx in group_cols)
		if not has_any_value:
			continue

		site_rows.setdefault(norm_site, []).append(row_idx)

	if not site_rows:
		ws._charts = []
		return 0

	for _, rows in site_rows.items():
		rows.sort(key=lambda r: sort_key_part(ws.cell(row=r, column=year_col).value))

	# Clear existing charts and rebuild from current data.
	ws._charts = []

	help_col = 60  # BH
	help_row = 1
	chart_index = 0
	for site_id in sorted(site_rows.keys(), key=lambda v: str(v).upper()):
		rows = site_rows[site_id]
		if not rows:
			continue

		ws.cell(row=help_row, column=help_col, value="year4")
		for i, (grp, _) in enumerate(group_cols, start=1):
			ws.cell(row=help_row, column=help_col + i, value=grp)

		for i, src_row in enumerate(rows, start=1):
			year_raw = ws.cell(row=src_row, column=year_col).value
			if year_raw is None or str(year_raw).strip() == "":
				year_label = None
			else:
				try:
					year_num = float(year_raw)
					year_label = str(int(year_num)) if year_num.is_integer() else str(year_raw)
				except Exception:
					year_label = str(year_raw)
			ws.cell(row=help_row + i, column=help_col, value=year_label)
			for j, (_, grp_col_idx) in enumerate(group_cols, start=1):
				ws.cell(row=help_row + i, column=help_col + j, value=ws.cell(row=src_row, column=grp_col_idx).value)

		data_ref = Reference(
			ws,
			min_col=help_col + 1,
			max_col=help_col + len(group_cols),
			min_row=help_row,
			max_row=help_row + len(rows),
		)
		cats_ref = Reference(
			ws,
			min_col=help_col,
			max_col=help_col,
			min_row=help_row + 1,
			max_row=help_row + len(rows),
		)

		chart = BarChart()
		chart.type = "col"
		chart.grouping = "stacked"
		chart.overlap = 100
		chart.title = f"Site {site_id}"
		chart.x_axis.delete = False
		chart.y_axis.delete = False
		chart.x_axis.axPos = "b"
		chart.y_axis.axPos = "l"
		chart.x_axis.tickLblPos = "nextTo"
		chart.y_axis.tickLblPos = "nextTo"
		chart.x_axis.tickLblSkip = 1
		chart.x_axis.tickMarkSkip = 1
		chart.x_axis.title = "year4"
		chart.y_axis.title = "value"
		chart.legend.position = "b"
		chart.legend.overlay = False
		chart.add_data(data_ref, titles_from_data=True)
		chart.set_categories(cats_ref)
		chart.height = 6.0
		chart.width = 8.5

		grid_col = 10 if chart_index % 2 == 0 else 19
		grid_row = 2 + (chart_index // 2) * 16
		anchor = f"{get_column_letter(grid_col)}{grid_row}"
		ws.add_chart(chart, anchor)

		chart_index += 1
		help_row += len(rows) + 3

	return chart_index


def build_lake_group_tables(long_df: pd.DataFrame, target_groups: list[str] = TARGET_GROUPS):
	if long_df.empty:
		raise ValueError("Long table is empty. Build long table first.")

	data = long_df.copy()
	data["group"] = data["group"].astype(str).str.strip()
	data["lakeid"] = data["lakeid"].astype(str).str.strip().str.upper()
	data["site"] = data["site"].apply(normalize_site_value)
	data["number_indiv"] = pd.to_numeric(data["number_indiv"], errors="coerce").fillna(0)
	data = data[data["site"].notna()].copy()

	target_lookup = {g.lower(): g for g in target_groups}
	data = data[data["group"].str.lower().isin(target_lookup.keys())].copy()
	data["group"] = data["group"].str.lower().map(target_lookup)

	# Sum taxa within each rep, then average those rep totals per site and group.
	rep_totals = (
		data.groupby(["lakeid", "year", "site", "rep", "group"], dropna=False, as_index=False)["number_indiv"]
		.sum()
	)
	site_group_avg = (
		rep_totals.groupby(["lakeid", "year", "site", "group"], dropna=False, as_index=False)["number_indiv"]
		.mean()
	)

	tables = {}
	for lake in ["SP", "CR", "TR"]:
		lake_df = site_group_avg[site_group_avg["lakeid"] == lake].copy()
		if lake_df.empty:
			tables[lake] = pd.DataFrame(columns=["year", "site", *target_groups])
			continue

		pivot = (
			lake_df.pivot_table(index=["year", "site"], columns="group", values="number_indiv", aggfunc="first")
			.reset_index()
		)
		for grp in target_groups:
			if grp not in pivot.columns:
				pivot[grp] = pd.NA
		pivot = pivot[["year", "site", *target_groups]]
		pivot = pivot.sort_values(
			by=["site", "year"],
			key=lambda s: s.map(lambda v: str(v).upper()) if s.name == "site" else s,
			kind="stable",
		).reset_index(drop=True)
		tables[lake] = pivot

	stats = {
		"input_rows": int(len(data)),
		"rep_rows": int(len(rep_totals)),
		"site_group_rows": int(len(site_group_avg)),
	}
	return tables, stats


def append_lake_group_tables_to_historical(hist_file_bytes: bytes, lake_tables: dict[str, pd.DataFrame]):
	def sort_sheet_by_site_year(ws):
		headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
		header_lookup = {}
		for idx, header in enumerate(headers, start=1):
			token = normalize_header_token(header)
			if token:
				header_lookup[token] = idx

		site_col = header_lookup.get("site")
		year_col = header_lookup.get("year4") or header_lookup.get("year")
		if site_col is None or year_col is None:
			return

		max_col = ws.max_column
		rows = []
		for row_idx in range(2, ws.max_row + 1):
			row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
			if any(v not in (None, "") for v in row_vals):
				rows.append(row_vals)

		rows.sort(
			key=lambda r: (
				str(normalize_site_value(r[site_col - 1]) or "").upper(),
				sort_key_part(r[year_col - 1]),
			)
		)

		for row_idx, row_vals in enumerate(rows, start=2):
			for col_idx, value in enumerate(row_vals, start=1):
				ws.cell(row=row_idx, column=col_idx, value=value)

	workbook = load_workbook(BytesIO(hist_file_bytes))
	sheet_map = {
		"SP": "SP grp",
		"CR": "CR grp",
		"TR": "TR grp",
	}
	ordered_cols = ["year", "site", "Amphipoda", "Diptera", "Ephemeroptera", "Gastropoda", "Odonata", "Oligochaeta"]
	start_col = 1

	append_summary = {}
	for lake, sheet_name in sheet_map.items():
		table = lake_tables.get(lake)
		if table is None or table.empty:
			append_summary[lake] = {"rows_appended": 0, "start_row": None, "end_row": None}
			continue

		ws = find_sheet_case_insensitive(workbook, sheet_name)
		if ws is None:
			raise ValueError(f"Sheet '{sheet_name}' not found in historical workbook.")

		data_cols = list(range(start_col, start_col + len(ordered_cols)))
		last_data_row = 1
		for row_idx in range(ws.max_row, 1, -1):
			if any(ws.cell(row=row_idx, column=col_idx).value not in (None, "") for col_idx in data_cols):
				last_data_row = row_idx
				break

		start_row = last_data_row + 1
		for _, row in table.iterrows():
			new_row = start_row
			for i, col_name in enumerate(ordered_cols):
				excel_col = start_col + i
				value = row[col_name]
				if pd.isna(value):
					value = None
				ws.cell(row=new_row, column=excel_col, value=value)
			start_row += 1

		append_summary[lake] = {
			"rows_appended": int(len(table)),
			"start_row": int(last_data_row + 1),
			"end_row": int(last_data_row + len(table)),
		}

	chart_counts = {}
	for lake, sheet_name in sheet_map.items():
		ws = find_sheet_case_insensitive(workbook, sheet_name)
		if ws is not None:
			sort_sheet_by_site_year(ws)
			chart_counts[lake] = regenerate_site_charts(ws, TARGET_GROUPS)
		else:
			chart_counts[lake] = 0

	buffer = BytesIO()
	workbook.save(buffer)
	buffer.seek(0)
	return buffer.getvalue(), append_summary, chart_counts


def build_final_historical_workbook(hist_file_bytes: bytes, long_df: pd.DataFrame, lake_tables: dict[str, pd.DataFrame]):
	# Step A: append full long table to historical "data" sheet.
	with_data_bytes, data_append_stats = append_long_table_to_historical(
		hist_file_bytes,
		long_df,
		sheet_name="data",
	)

	# Step B: append lake aggregate tables, sort, and regenerate charts.
	final_bytes, grp_append_stats, grp_chart_stats = append_lake_group_tables_to_historical(
		with_data_bytes,
		lake_tables,
	)

	return final_bytes, data_append_stats, grp_append_stats, grp_chart_stats


def main() -> None:
	st.sidebar.header("Input Files")
	raw_file = st.sidebar.file_uploader(
		"Raw Dendy Data (.xlsx)",
		type=["xlsx"],
		help="Upload the raw dendy export workbook.",
	)
	hist_file = st.sidebar.file_uploader(
		"Historical Dendy Data (.xlsx) - optional for now",
		type=["xlsx"],
		help="Upload the historical dendy workbook.",
	)

	if raw_file is None:
		st.info("Upload Raw Dendy Data to continue.")
		return

	try:
		raw_df = load_first_sheet(raw_file)
		raw_matrix = load_sheet_matrix(raw_file)
		st.session_state["raw_dendy_df"] = raw_df
		st.session_state["raw_dendy_matrix"] = raw_matrix
		if hist_file is not None:
			hist_file.seek(0)
			hist_file_bytes = hist_file.getvalue()
			hist_workbook, hist_sheets = load_workbook_bundle(hist_file_bytes)
			st.session_state["historical_dendy_bytes"] = hist_file_bytes
			st.session_state["historical_dendy_workbook"] = hist_workbook
			st.session_state["historical_dendy_sheets"] = hist_sheets
		else:
			hist_sheets = None
	except Exception as exc:
		st.error(f"Excel load error: {exc}")
		return

	c1, c2 = st.columns(2)
	with c1:
		st.subheader("Raw Dendy Data")
		st.success(f"Loaded {len(raw_df):,} rows x {len(raw_df.columns)} columns")
		st.dataframe(raw_df.head(15), use_container_width=True, hide_index=True)

	with c2:
		st.subheader("Historical Dendy Data")
		if hist_sheets is None:
			st.info("Historical file is optional in this step. You can add it later for row-bind.")
		else:
			hist_sheet_names = list(hist_sheets.keys())
			selected_sheet = st.selectbox("Choose a historical sheet", hist_sheet_names, key="historical_sheet_choice")
			hist_df = hist_sheets[selected_sheet]
			st.success(
				f"Loaded {len(hist_sheet_names):,} sheet(s); previewing '{selected_sheet}' with "
				f"{len(hist_df):,} rows x {len(hist_df.columns)} columns"
			)
			st.dataframe(hist_df.head(15), use_container_width=True, hide_index=True)

	st.divider()
	st.subheader("Step 1: Data Manager Export")
	st.write(
		"Build a new XLSX from the raw file using the standard Dendy layout. "
		"Taxa with zero counts across all location columns are dropped."
	)
	export_year = st.number_input(
		"Export year",
		min_value=1900,
		max_value=2100,
		value=2025,
		step=1,
		help="Year written to the export file for every taxon row.",
	)
	data_start_col = st.number_input(
		"Raw data start column (1-based)",
		min_value=1,
		max_value=500,
		value=12,
		step=1,
		help="Per your format, count data starts at column 12.",
	)

	if st.button("Create Data Manager Export", type="primary"):
		try:
			export_bytes, export_preview_df, export_preview_matrix_df, export_stats = build_data_manager_export(
				raw_matrix,
				export_year=int(export_year),
				data_start_col_1_based=int(data_start_col),
			)
			st.session_state["dendy_export_bytes"] = export_bytes
			st.session_state["dendy_export_preview_df"] = export_preview_df
			st.session_state["dendy_export_preview_matrix_df"] = export_preview_matrix_df
			st.session_state["dendy_export_stats"] = export_stats
		except Exception as exc:
			st.error(f"Export build error: {exc}")

	export_stats = st.session_state.get("dendy_export_stats")
	export_preview_df = st.session_state.get("dendy_export_preview_df")
	export_preview_matrix_df = st.session_state.get("dendy_export_preview_matrix_df")
	export_bytes = st.session_state.get("dendy_export_bytes")

	if export_stats and export_preview_df is not None and export_preview_matrix_df is not None and export_bytes:
		m1, m2, m3, m4 = st.columns(4)
		m1.metric("Taxa in raw", export_stats["input_taxa"])
		m2.metric("Taxa kept", export_stats["kept_taxa"])
		m3.metric("Taxa dropped", export_stats["dropped_taxa"])
		m4.metric("Location columns", export_stats["location_columns"])

		st.write("Export layout preview")
		st.dataframe(export_preview_matrix_df.head(28), use_container_width=True, hide_index=True)
		st.download_button(
			"Download Data Manager Export (.xlsx)",
			data=export_bytes,
			file_name="dendy_data_manager_export.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	st.divider()
	st.subheader("Step 2: Long Table Export")
	st.write(
		"Create a long-format table for Data Manager with headers: "
		"lakeid, year, site, rep, group, family, taxon_code, description, number_indiv."
	)
	if st.button("Create Long Table", type="secondary"):
		try:
			long_df, long_bytes, long_stats = build_long_table_export(
				raw_matrix,
				export_year=int(export_year),
				data_start_col_1_based=int(data_start_col),
			)
			st.session_state["dendy_long_df"] = long_df
			st.session_state["dendy_long_bytes"] = long_bytes
			st.session_state["dendy_long_stats"] = long_stats
		except Exception as exc:
			st.error(f"Long table build error: {exc}")

	long_df = st.session_state.get("dendy_long_df")
	long_bytes = st.session_state.get("dendy_long_bytes")
	long_stats = st.session_state.get("dendy_long_stats")
	if long_df is not None and long_bytes is not None and long_stats is not None:
		l1, l2, l3 = st.columns(3)
		l1.metric("Long rows", long_stats["rows"])
		l2.metric("Taxa", long_stats["taxa"])
		l3.metric("Lakes", long_stats["lakes"])
		st.dataframe(long_df.head(30), use_container_width=True, hide_index=True)
		st.download_button(
			"Download Long Table (.xlsx)",
			data=long_bytes,
			file_name="dendy_long_table.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	st.divider()
	st.subheader("Step 3: Append Long Table to Historical")
	st.write("Append the long table to the bottom of the historical workbook sheet named 'data'.")

	hist_file_bytes = st.session_state.get("historical_dendy_bytes")
	if hist_file_bytes is None:
		st.info("Upload Historical Dendy Data to enable append.")
	elif long_df is None:
		st.info("Create Long Table first, then append to historical workbook.")
	else:
		if st.button("Append to Historical 'data' Sheet", type="secondary"):
			try:
				updated_hist_bytes, append_stats = append_long_table_to_historical(
					hist_file_bytes,
					long_df,
					sheet_name="data",
				)
				st.session_state["updated_historical_bytes"] = updated_hist_bytes
				st.session_state["append_stats"] = append_stats
			except Exception as exc:
				st.error(f"Append error: {exc}")

	append_stats = st.session_state.get("append_stats")
	updated_hist_bytes = st.session_state.get("updated_historical_bytes")
	if append_stats is not None and updated_hist_bytes is not None:
		a1, a2, a3 = st.columns(3)
		a1.metric("Rows appended", append_stats["rows_appended"])
		a2.metric("Start row", append_stats["start_row"])
		a3.metric("End row", append_stats["end_row"])
		st.download_button(
			"Download Updated Historical Workbook (.xlsx)",
			data=updated_hist_bytes,
			file_name="historical_dendy_updated.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	st.divider()
	st.subheader("Step 4: Lake Aggregate Group Tables")
	st.write(
		"Create lake-specific aggregate tables for groups Amphipoda, Diptera, Ephemeroptera, "
		"Gastropoda, Odonata, and Oligochaeta."
	)
	if long_df is None:
		st.info("Create Long Table first to build aggregates.")
	else:
		if st.button("Create Lake Aggregate Tables", type="secondary"):
			try:
				lake_tables, lake_stats = build_lake_group_tables(long_df, TARGET_GROUPS)
				st.session_state["lake_group_tables"] = lake_tables
				st.session_state["lake_group_stats"] = lake_stats
			except Exception as exc:
				st.error(f"Aggregate table error: {exc}")

	lake_tables = st.session_state.get("lake_group_tables")
	lake_stats = st.session_state.get("lake_group_stats")
	if lake_tables is not None and lake_stats is not None:
		s1, s2, s3 = st.columns(3)
		s1.metric("Filtered long rows", lake_stats["input_rows"])
		s2.metric("Rep-level rows", lake_stats["rep_rows"])
		s3.metric("Site-group averages", lake_stats["site_group_rows"])

		tab_sp, tab_cr, tab_tr = st.tabs(["SP", "CR", "TR"])
		with tab_sp:
			st.dataframe(lake_tables["SP"].head(40), use_container_width=True, hide_index=True)
		with tab_cr:
			st.dataframe(lake_tables["CR"].head(40), use_container_width=True, hide_index=True)
		with tab_tr:
			st.dataframe(lake_tables["TR"].head(40), use_container_width=True, hide_index=True)

		if hist_file_bytes is None:
			st.info("Upload Historical Dendy Data to append aggregate tables.")
		else:
			if st.button("Append Aggregate Tables to SP grp / CR grp / TR grp", type="secondary"):
				try:
					updated_grp_bytes, grp_append_stats, grp_chart_stats = append_lake_group_tables_to_historical(hist_file_bytes, lake_tables)
					st.session_state["updated_historical_grp_bytes"] = updated_grp_bytes
					st.session_state["grp_append_stats"] = grp_append_stats
					st.session_state["grp_chart_stats"] = grp_chart_stats
				except Exception as exc:
					st.error(f"Aggregate append error: {exc}")

	grp_append_stats = st.session_state.get("grp_append_stats")
	grp_chart_stats = st.session_state.get("grp_chart_stats")
	updated_grp_bytes = st.session_state.get("updated_historical_grp_bytes")
	if grp_append_stats is not None and updated_grp_bytes is not None:
		st.write("Aggregate append summary")
		summary_rows = []
		for lake in ["SP", "CR", "TR"]:
			r = grp_append_stats.get(lake, {})
			summary_rows.append(
				{
					"lakeid": lake,
					"rows_appended": r.get("rows_appended"),
					"start_row": r.get("start_row"),
					"end_row": r.get("end_row"),
					"charts_created": grp_chart_stats.get(lake) if isinstance(grp_chart_stats, dict) else None,
				}
			)
		st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)
		st.download_button(
			"Download Historical Workbook with Group Aggregates (.xlsx)",
			data=updated_grp_bytes,
			file_name="historical_dendy_with_group_aggregates.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	st.divider()
	st.subheader("Final Output: All Steps in One Workbook")
	st.write("Build one historical workbook that includes: long-table append to 'data', and SP/CR/TR group updates + charts.")
	if hist_file_bytes is None:
		st.info("Upload Historical Dendy Data to build final workbook.")
	elif long_df is None:
		st.info("Create Long Table first to build final workbook.")
	elif lake_tables is None:
		st.info("Create Lake Aggregate Tables first to build final workbook.")
	else:
		if st.button("Build Final Historical Workbook", type="primary"):
			try:
				final_bytes, final_data_stats, final_grp_stats, final_chart_stats = build_final_historical_workbook(
					hist_file_bytes,
					long_df,
					lake_tables,
				)
				st.session_state["final_historical_bytes"] = final_bytes
				st.session_state["final_data_append_stats"] = final_data_stats
				st.session_state["final_grp_append_stats"] = final_grp_stats
				st.session_state["final_chart_stats"] = final_chart_stats
			except Exception as exc:
				st.error(f"Final build error: {exc}")

	final_bytes = st.session_state.get("final_historical_bytes")
	final_data_stats = st.session_state.get("final_data_append_stats")
	final_grp_stats = st.session_state.get("final_grp_append_stats")
	final_chart_stats = st.session_state.get("final_chart_stats")
	if final_bytes is not None and final_data_stats is not None and final_grp_stats is not None:
		st.write("Final build summary")
		f1, f2, f3 = st.columns(3)
		f1.metric("Data rows appended", final_data_stats.get("rows_appended"))
		f2.metric("Data start row", final_data_stats.get("start_row"))
		f3.metric("Data end row", final_data_stats.get("end_row"))

		summary_rows = []
		for lake in ["SP", "CR", "TR"]:
			r = final_grp_stats.get(lake, {})
			summary_rows.append(
				{
					"lakeid": lake,
					"rows_appended": r.get("rows_appended"),
					"start_row": r.get("start_row"),
					"end_row": r.get("end_row"),
					"charts_created": final_chart_stats.get(lake) if isinstance(final_chart_stats, dict) else None,
				}
			)
		st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

		st.download_button(
			"Download Final Historical Workbook (.xlsx)",
			data=final_bytes,
			file_name="historical_dendy_final_all_updates.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	st.info(
		"Historical workbook is loaded with all sheets available in memory. "
		"If we write back with openpyxl, the workbook object can keep charts and other embedded objects."
	)

	st.caption("Next step: append (row-bind) processed records into the historical workbook.")


if __name__ == "__main__":
	main()
